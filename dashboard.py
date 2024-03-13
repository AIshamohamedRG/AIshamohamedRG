from tkinter import *
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

def update_dashboard():
    
    wb = load_workbook('aisha/supermarkt_sales.xlsx')
    ws = wb.active
    
    column_Q = ws['Q']
    gross_income = [cell.value if isinstance(cell.value, (int, float)) else 0 for cell in column_Q]

    
    plt.clf()
    
    plt.plot(range(1, len(gross_income) + 1), gross_income)
    plt.xlabel('Transaction Number')
    plt.ylabel('Gross Income')
    plt.title('Gross Income Over Transactions')
    
    
    plot_canvas.draw()

    
    column_K = ws['K']
    total_sales = sum(cell.value if isinstance(cell.value, (int, float)) else 0 for cell in column_K)

    column_R = ws['R']
    average_rating = sum(cell.value if isinstance(cell.value, (int, float)) else 0 for cell in column_R) / len(column_R)

    column_I = ws['I']
    total_transactions = sum(1 for _ in column_I)
    average_sale_per_transaction = total_sales / total_transactions

    
    lbl_total_sales.config(text='Total Sales: {:.2f}'.format(total_sales))
    lbl_average_rating.config(text='Average Rating: {:.2f}'.format(average_rating))
    lbl_average_sale_per_transaction.config(text='Average Sales Per Transaction: {:.2f}'.format(average_sale_per_transaction))


aisha = Tk()
aisha.geometry('900x600')
aisha.title('Sales Dashboard')


wb = load_workbook('aisha/supermarkt_sales.xlsx')
ws = wb.active


lbl_total_sales = Label(aisha, text='Total Sales: N/A', font=('Arial', 18), bg='gray')
lbl_total_sales.grid(row=0, column=0, padx=10, pady=10)

lbl_average_rating = Label(aisha, text='Average Rating: N/A', font=('Arial', 18), bg='gray')
lbl_average_rating.grid(row=0, column=1, padx=10, pady=10)

lbl_average_sale_per_transaction = Label(aisha, text='Average Sales Per Transaction: N/A', font=('Arial', 18), bg='gray')
lbl_average_sale_per_transaction.grid(row=0, column=2, padx=10, pady=10)


plot_frame = Frame(aisha)
plot_frame.grid(row=1, column=0, columnspan=3, padx=10, pady=10)
plt.figure(figsize=(6, 4))
plot_canvas = FigureCanvasTkAgg(plt.gcf(), master=plot_frame)
plot_canvas.get_tk_widget().pack()


update_dashboard()

aisha.mainloop()








